import json
import os
import urllib.request
import re
from datetime import datetime, timedelta
import openpyxl

DATA_FILE = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'calendario.json')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'Calendario_SerieA_2026-27.xlsx')

WC_API_URL = "https://raw.githubusercontent.com/openfootball/world-cup.json/master/2026/worldcup.json"

# ==========================================
# 1. MONDIALI 2026 (DA INTERNET JSON)
# ==========================================
def converti_orario_messicano_in_italiano(data_str, time_str):
    """
    Legge orari come "13:00 UTC-6", aggiunge 8 ore (differenza con Italia)
    e formatta la data. Gestisce il superamento della mezzanotte automaticamente.
    """
    try:
        orario_base = time_str.split(' ')[0]
        ore, minuti = map(int, orario_base.split(':'))
        fuso_str = time_str.split('UTC')[1]
        fuso_originale = int(fuso_str) if fuso_str else 0
        
        # Italia (UTC+2) - Fuso Messico (es. -6) = +8 ore da aggiungere
        differenza_ore = 2 - fuso_originale
        
        anno, mese, giorno = map(int, data_str.split('-'))
        dt = datetime(anno, mese, giorno, ore, minuti)
        dt_italiano = dt + timedelta(hours=differenza_ore)
        return dt_italiano.strftime("%Y-%m-%dT%H:%M:00+02:00")
    except Exception as e:
        return f"{data_str}T15:00:00+02:00" # Fallback generico

def estrai_mondiali():
    print("Scaricando Mondiali 2026 da Internet...")
    req = urllib.request.Request(WC_API_URL, headers={'User-Agent': 'Mozilla/5.0'})
    try:
        with urllib.request.urlopen(req) as response:
            dati = json.loads(response.read().decode())
    except Exception as e:
        print(f"Errore download Mondiali: {e}")
        return []
        
    eventi = []
    for partita in dati.get('matches', []):
        data_match = partita.get('date')
        time_match = partita.get('time', '00:00 UTC+0')
        orario_italiano_iso = converti_orario_messicano_in_italiano(data_match, time_match)
        
        squadra1 = partita.get('team1', 'TBD')
        squadra2 = partita.get('team2', 'TBD')
        fase = partita.get('round', 'Fase a Gironi')
        
        id_partita = f"wc26_{data_match}_{squadra1}_{squadra2}".replace(" ", "_").lower()
        
        eventi.append({
            "id": id_partita,
            "sport": "Calcio",
            "competition": "Mondiali 2026",
            "eventName": f"{squadra1} vs {squadra2} ({fase})",
            "dateTime": orario_italiano_iso,
            "broadcaster": "Rai / Sky"
        })
    return eventi

# ==========================================
# 2. SERIE A (DAL FILE EXCEL)
# ==========================================
def estrai_serie_a_inter_da_excel():
    print("Estraendo calendario Serie A (Solo Inter) dal file Excel...")
    if not os.path.exists(EXCEL_FILE):
        print(f"Errore: File Excel non trovato: {EXCEL_FILE}")
        return []

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        print(f"Errore caricamento Excel: {e}")
        return []

    eventi = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Scansioniamo tutte le celle
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row):
                if not cell_value or not isinstance(cell_value, str):
                    continue
                
                # Se la cella contiene la parola "Inter" (indipendentemente da maiuscole/minuscole)
                # ed ha un trattino in mezzo (es. Bologna-Inter)
                if "inter" in cell_value.lower() and "-" in cell_value:
                    partita_nome = cell_value.strip()
                    
                    # Caccia alla data: risaliamo le righe in alto nella stessa colonna
                    data_trovata = None
                    for r_up in range(row_idx - 1, 0, -1):
                        cella_sopra = ws.cell(row=r_up, column=col_idx+1).value
                        if cella_sopra and isinstance(cella_sopra, str) and "(" in cella_sopra and ")" in cella_sopra:
                            # Cerchiamo un formato come 23/08/2026 o 06/09/26 tra parentesi
                            match = re.search(r'\((\d{2}/\d{2}/\d{2,4})\)', cella_sopra)
                            if match:
                                data_trovata = match.group(1)
                                break
                    
                    if data_trovata:
                        giorno, mese, anno = data_trovata.split('/')
                        if len(anno) == 2:
                            anno = f"20{anno}"
                        iso_date = f"{anno}-{mese}-{giorno}T15:00:00+02:00"
                        # Aggiungiamo un hash per garantire che ogni partita abbia una chiave univoca per React!
                        id_partita = f"seriea_{anno}{mese}{giorno}_inter_{abs(hash(partita_nome))}"
                        
                        eventi.append({
                            "id": id_partita,
                            "sport": "Calcio",
                            "competition": "Serie A 26/27",
                            "eventName": partita_nome,
                            "dateTime": iso_date,
                            "broadcaster": "DAZN / Sky"
                        })
    return eventi

# ==========================================
# 3. CHAMPIONS LEAGUE (IN STANDBY)
# ==========================================
def estrai_champions_league():
    # DA ATTIVARE A FINE AGOSTO 2026
    print("Champions League in standby. (Nessun download effettuato).")
    return []

# ==========================================
# MOTORE DI SALVATAGGIO 
# ==========================================
def salva_nel_database_locale(nuovi_eventi):
    if not nuovi_eventi:
        return

    dati_esistenti = []
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as file:
            try:
                dati_esistenti = json.load(file)
            except json.JSONDecodeError:
                pass
            
    # Manteniamo intatta la F1 che abbiamo caricato prima, ma purifichiamo il calcio vecchio
    dati_puliti = [evento for evento in dati_esistenti if evento.get('sport') != 'Calcio']
    
    dati_finali = dati_puliti + nuovi_eventi
    
    with open(DATA_FILE, 'w', encoding='utf-8') as file:
        json.dump(dati_finali, file, indent=2, ensure_ascii=False)
        
    print(f"Lavoro completato! Inseriti {len(nuovi_eventi)} eventi calcistici nel database.")

# Esecuzione del Bot
if __name__ == "__main__":
    # 1. Scarica da Internet
    eventi_mondiali = estrai_mondiali()
    # 2. Legge Excel Locale
    eventi_serie_a = estrai_serie_a_inter_da_excel()
    # 3. Standby
    # eventi_champions = estrai_champions_league()
    
    tutto_il_calcio = eventi_mondiali + eventi_serie_a
    salva_nel_database_locale(tutto_il_calcio)
